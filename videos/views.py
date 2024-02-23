import openpyxl
import json
from django.http import HttpResponse
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from .models import VideoResponse

@csrf_exempt
@require_http_methods(['POST'])
def crear_video_response(request):
    try:
        # Decodifica request.body a str y luego procesa el JSON
        data = json.loads(request.body.decode('utf-8'))
        
        # Aquí puedes usar 'data' para crear un nuevo objeto VideoResponse
        video_response = VideoResponse.objects.create(
            matricula=data['matricula'],
            video_name=data['video_name'],
            response=data['response']
        )
        
        # Envía una respuesta JSON de éxito
        return JsonResponse({'mensaje': 'VideoResponse creado con éxito', 'id': video_response.id}, status=201)
    except json.JSONDecodeError:
        # Maneja el caso en que el cuerpo de la solicitud no es JSON válido
        return JsonResponse({'error': 'Formato JSON inválido'}, status=400)
    except KeyError:
        # Maneja el caso en que falta alguna clave en los datos JSON
        return JsonResponse({'error': 'Falta una clave requerida en los datos JSON'}, status=400)

def video_list(request):
    videos = VideoResponse.objects.all()
    videos_list = ", ".join([video.video_name for video in videos])
    return HttpResponse(f"Lista de Videos: {videos_list}")

def video_detail(request, video_id):
    video = VideoResponse.objects.get(id=video_id)
    return HttpResponse(f"Video: {video.video_name}, Respuesta: {video.response}")

def export_videos_to_excel(request):
    # Crea un libro de trabajo y una hoja
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Videos y Respuestas'

    # Añade títulos de columna en la primera fila, excluyendo 'ID'
    columns = ['ID Usuario', 'Nombre del Video', 'Respuesta']
    for col_num, column_title in enumerate(columns, start=1):
        worksheet.cell(row=1, column=col_num, value=column_title)

    # Itera sobre los videos y respuestas para llenar las celdas, excluyendo 'ID'
    for row_num, video in enumerate(VideoResponse.objects.all(), start=2):
        row_data = [
            video.matricula,
            video.video_name,
            video.response,
        ]
        
        for col_num, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Configura la respuesta para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="videos-respuestas.xlsx"'

    # Guarda el libro de trabajo en la respuesta
    workbook.save(response)

    return response